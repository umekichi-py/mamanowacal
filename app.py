import os, calendar, re
from datetime import datetime
from flask import Flask, request, render_template, redirect, url_for, session, flash, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_LEFT
import io
from user_repository import UserRepository
from calendar_repository import CalendarRepository

repo = UserRepository()
calendar_repo = CalendarRepository()

font_path = os.path.join(os.path.dirname(__file__), "fonts", "NotoSansJP-Regular.ttf")
if os.path.exists(font_path):
    pdfmetrics.registerFont(TTFont('NotoSansJP', font_path))
    PDF_JP_FONT = 'NotoSansJP'
else:
    windows_font = os.path.join(os.getenv('WINDIR', 'C:\\Windows'), 'Fonts', 'msgothic.ttc')
    if os.path.exists(windows_font):
        pdfmetrics.registerFont(TTFont('MSGOTHIC', windows_font))
        PDF_JP_FONT = 'MSGOTHIC'
    else:
        PDF_JP_FONT = 'Helvetica'

#admin作成
def init_admin():

    if repo.get_user("admin") is None:

        repo.create_user(
            username="admin",
            password=generate_password_hash("mamanowa"),
            role="admin"
        )
'''
    users = load_users()

    if "admin" not in users:
        users["admin"] = {
            "password": generate_password_hash("admin123"),
            "role": "admin"
        }
        save_users(users)
        print("管理者アカウント作成:admin / admin123")
'''

#イベントファイル読み込みの共通関数
def load_events(username, mode):
    return calendar_repo.get_all_events(username, mode)

def load_users():
    return repo.get_all_users()


def get_user_display_name(username, user, mode=None):
    user_data = user or {}
    if mode == "childday":
        return user_data.get("child_name") or ""
    return f"{user_data.get('job') or ''}{user_data.get('staff_id') or ''}{username}"


def sort_users_for_display(users):
    job_order = {"育": 1, "援": 2, "給": 3, "看": 4}
    return sorted(
        users.items(),
        key=lambda item: (
            job_order.get((item[1] or {}).get("job") or "", 99),
            (item[1] or {}).get("staff_id") or "",
            item[0]
        )
    )

def get_display_name(username, user, mode):
    user_data = user or {}
    if mode == "childday":
        return user_data.get("child_name") or ""

    return f"{user_data.get('job') or ''}{user_data.get('staff_id') or ''}{username}"


app: Flask = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(24))
init_admin()

#ユーザー登録機能
@app.route("/register", methods=["GET", "POST"])
def register():

#admin以外はとうろくできないように
    if "user" not in session:
        return redirect(url_for("login"))
    
    if session.get("role") != "admin":
        flash("管理者のみ登録できます。", "error")
        return redirect(url_for("home"))

    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        role = request.form.get("role", "user")
        staff_id = request.form.get("staff_id")
        job = request.form.get("job")
        child_name = request.form.get("child_name")

        if not username or not password:
            flash("未入力があります。", "error")
            return render_template("register.html")
        
        if repo.get_user(username):
            flash("登録済みです", "error")
            return render_template("register.html")

        """
        users = load_users()

        if username in users:
            flash("登録済みです。", "error")
            return render_template("register.html")
        """

        hashed_pw = generate_password_hash(password)

        repo.create_user(
            username=username,
            password=hashed_pw,
            role=role,
            staff_id=staff_id,
            job=job,
            child_name=child_name
        )

        '''
        users[username] = {
            "password": hashed_pw,
            "role": role,
            "staff_id": staff_id,
            "job": job,
            "child_name": child_name
        }
        '''

        flash("ユーザー登録が完了しました。", "success")
        return render_template("register.html")
    
    return render_template("register.html")

#ログインページ
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        
        users = load_users()
    
        if username in users and check_password_hash(
            users[username]["password"], password
        ):
            session["user"] = username
            session["role"] = users[username]["role"]
            return redirect(url_for("home"))
        
        flash("ユーザー名またはパスワードが違います。", "error")
        return render_template("login.html")
        
    return render_template("login.html")

#ログアウト
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

#パスワードの変更機能
@app.route("/change_password", methods=["GET", "POST"])
def change_password():

    if "user" not in session:
        return redirect(url_for("login"))
    
    username = session["user"]

    if request.method == "POST":

        current_pw = request.form.get("current_pw")
        new_pw = request.form.get("new_pw")
        confirm_pw = request.form.get("confirm_pw")

        users = load_users()

        #現在パスワード確認
        if not check_password_hash(users[username]["password"], current_pw):
            flash("現在のパスワードが違います。", "danger")
            return redirect(url_for("change_password"))

        if len(new_pw) < 8:
            flash("パスワードは8文字以上にしてください。")
            return redirect(url_for("change_password"))
        
        if new_pw != confirm_pw:
            flash("新しいパスワードが一致しません。", "danger")
            return redirect(url_for("change_password"))
        
        #更新
        repo.update_user(username, password=generate_password_hash(new_pw))

        flash("パスワードを変更しました。", "success")
        return redirect(url_for("change_password"))
    
    return render_template("change_password.html")

#管理者ページ
@app.route("/admin")
def admin_page():

    if "user" not in session:
        return redirect(url_for("login"))

    if session.get("role") != "admin":
        return "不正な操作です。ページを開きなおしてください。"
    
    users = load_users()

    return render_template("admin.html", users=users)

#ユーザー情報編集ページ
@app.route("/admin/edit_user/<username>", methods=["GET", "POST"])
def edit_user(username):

    if "user" not in session:
        return redirect(url_for("login"))
    
    if session.get("role") != "admin":
        return "不正な操作です"

    users = load_users()

    if username not in users:
        return "ユーザーが存在しません"

    # POST（更新処理）
    if request.method == "POST":
        staff_id = request.form.get("staff_id")
        job = request.form.get("job")
        child_name = request.form.get("child_name")
        role = request.form.get("role")

        if username == "admin" and role != "admin":
            flash("adminの権限は変更できません", "danger")
            return redirect(url_for("admin_page", username=username))

        repo.update_user(
            username,
            role=role,
            staff_id=staff_id,
            job=job,
            child_name=child_name
        )

        flash("更新しました", "success")
        return redirect(url_for("admin_page"))

    # GET（表示）
    user = users[username]

    return render_template(
        "edit_user.html",
        username=username,
        user=user
    )

#パスワードのリセット
@app.route("/admin/reset/<username>", methods=["GET", "POST"])
def reset(username):

    if "user" not in session:
        return redirect(url_for("login"))
    
    if session.get("role") != "admin":
        return "不正な操作です。ページを開きなおしてください。"
    
    users = load_users()

    if username not in users:
        flash("ユーザーが存在しません。", "danger")
        return redirect(url_for("reset", username=username))
    
    if request.method == "POST":

        temp_pw = request.form.get("temp_pw")

        if len(temp_pw) < 8:
            flash("パスワードは8文字以上にしてください。", "danger")
            return redirect(url_for("reset", username=username))
        
        repo.update_user(username, password=generate_password_hash(temp_pw))

        flash(f"{username}のパスワードを変更しました。", "success")

        return redirect(url_for("reset", username=username))
    
    return render_template("reset.html", username=username)

#ユーザー削除
@app.route("/admin/delete_user/<username>", methods=["POST"])
def delete_user(username):

    if "user" not in session:
        return redirect(url_for("login"))
    
    if session.get("role") != "admin":
        return "不正な操作です。ページを開きなおしてください。"
    
    users = load_users()

    #admin削除防止
    if username == "admin":
        flash("adminは削除できません。", "danger")
        return redirect(url_for("admin_page"))
    
    if username in users:
        repo.delete_user(username)
        calendar_repo.delete_user_events(username)
        flash(f"{username}を削除しました。", "success")

    return redirect(url_for('admin_page'))

#ロール切り替え
@app.route("/admin/toggle_role/<username>", methods=["POST"])
def toggle_role(username):

    if "user" not in session:
        return redirect(url_for("login"))
    
    if session.get("role") != "admin":
        return "不正な操作です。ページを開きなおしてください。"
    
    users = load_users()

    if username not in users:
        flash("ユーザーが存在しません。", "danger")
        return redirect(url_for("admin_page"))
    
    #adminのロール変更防止
    if username == "admin":
        flash("adminの属性は変更できません。", "danger")
        return redirect(url_for("admin_page"))
    
    current_role = users[username]["role"]

    role_order = ["user", "withchild", "childonly", "admin"]

    index = role_order.index(current_role)
    new_role = role_order[(index + 1) % len(role_order)]

    repo.update_user(username, role=new_role)

    flash(f"{username}の属性を{new_role}に変更しました。", "success")
    return redirect(url_for("admin_page"))

#一覧カレンダーを選ぶ
@app.route("/admin/calendar")
def admin_calendar():

    if "user" not in session:
        return redirect(url_for("login"))
    
    if session.get("role") != "admin":
        return "不正な操作です。ページを開きなおしてください。"
    
    now = datetime.now()
    default_month = now.strftime("%Y-%m")

    return render_template(
        "admin_calendar_select.html",
        default_month = default_month
        )

#月一覧を作る
@app.route("/admin/calendar/view")
def admin_calendar_view():

    if session.get("role") != "admin":
        return redirect(url_for("login"))
    
    mode = request.args.get("mode")
    month_str = request.args.get("month")

    if not month_str:
        now = datetime.now()
        year = now.year
        month = now.month
    else:
        year, month = map(int, month_str.split("-"))

    days = calendar.monthrange(year, month)[1]

    users = load_users()
    users_sorted = sort_users_for_display(users)
    usernames = [username for username, _ in users_sorted]
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{days:02d}"
    all_events = calendar_repo.get_events_by_users(usernames, mode, start_date, end_date)

    table = {}

    for day in range(1, days +1):
        date = f"{year}-{month:02d}-{day:02d}"
        table[date] = {}

        for username, user in users_sorted:
            table[date][username] = all_events.get(username, {}).get(date, {})

    return render_template(
        "admin_calendar_table.html",
        table = table,
        users = users,
        users_sorted=users_sorted,
        year = year,
        month = month,
        mode = mode,
        get_display_name=get_display_name
    )

#Excel出力
@app.route("/admin/calendar/export")
def export_calendar():

    def format_time(t):
        if not t:
            return ""
        h, m = t.split(":")
        return f"{int(h)}{m}"

    if session.get("role") != "admin":
        return redirect(url_for("login"))
    
    mode = request.args.get("mode")
    month_str = request.args.get("month")

    if not month_str:
        now = datetime.now()
        year = now.year
        month = now.month

    else:
        year, month = map(int, month_str.split("-"))

    users = load_users()
    users_sorted = sort_users_for_display(users)
    days = calendar.monthrange(year, month)[1]

    usernames = [username for username, _ in users_sorted]
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{days:02d}"
    all_events = calendar_repo.get_events_by_users(usernames, mode, start_date, end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month}"

    #ヘッダー行
    ws.cell(row=1, column=1, value="日付")

    col = 2
    for username, user in users_sorted:
        display_name = get_display_name(username, user, mode)
        ws.cell(row=1, column=col, value=display_name)
        ws.cell(row=1, column=col+1, value="")

        #1ユーザー2列使用
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=1, end_column=col+1)
        col += 2

    #データ
    for day in range(1, days + 1):

        row1 = 2 + (day - 1) * 2 #上段
        row2 = row1 + 1 #下段

        date = f"{year}-{month:02d}-{day:02d}"
        ws.cell(row=row1, column=1, value=date)

        col = 2
        for username, _ in users_sorted:
            data = all_events.get(username, {})

            time_S = ""
            time_E = ""
            comment = ""

            if date in data:
                e = data[date]
                time_S = e.get('timeS', '')
                time_E = e.get('timeE', '')
                comment = e.get("comment", "")
            
            #上段：開始時刻+コメント
            ws.cell(row=row1, column=col, value=format_time(time_S))
            ws.cell(row=row1, column=col+1, value=comment)

            #下段終了時刻
            ws.cell(row=row2, column=col, value=format_time(time_E))
            
            col += 2

    #メモリに保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"calendar_{mode}_{year}_{month}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/admin/calendar/export_pdf")
def export_calendar_pdf():
    if session.get("role") != "admin":
        return redirect(url_for("login"))

    mode = request.args.get("mode")
    month_str = request.args.get("month")

    if not month_str:
        now = datetime.now()
        year = now.year
        month = now.month
    else:
        year, month = map(int, month_str.split("-"))

    users = load_users()
    users_sorted = sort_users_for_display(users)
    days = calendar.monthrange(year, month)[1]

    usernames = [username for username, _ in users_sorted]
    start_date = f"{year}-{month:02d}-01"
    end_date = f"{year}-{month:02d}-{days:02d}"
    all_events = calendar_repo.get_events_by_users(usernames, mode, start_date, end_date)

    title_map = {
        "holiday": "休み希望用",
        "workday": "出勤希望用",
        "childday": "子ども預け希望"
    }
    title = title_map.get(mode, "カレンダー")

    story = []
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.fontName = PDF_JP_FONT
    title_style.fontSize = 14
    title_style.leading = 16
    story.append(Paragraph(f"{year}年{month}月 {title}", title_style))
    story.append(Spacer(1, 6))

    headers = ["日付"] + [get_user_display_name(username, user, mode) for username, user in users_sorted]
    rows = [headers]

    for day in range(1, days + 1):
        date = f"{year}-{month:02d}-{day:02d}"
        row = [date]
        for username, _ in users_sorted:
            data = all_events.get(username, {}).get(date, {})
            if data:
                timeS = data.get("timeS", "")
                timeE = data.get("timeE", "")
                comment = data.get("comment", "")

                text = f"{timeS}-{timeE}"

                if comment:
                    text += f"{comment}"
                row.append(text)
            else:
                row.append("")
        rows.append(row)

    page_width, page_height = A4
    MAX_USERS = 5

    fixed_header = headers[0]
    user_headers = headers[1:]

    for col_start in range(0, len(user_headers), MAX_USERS):

        page_headers = [fixed_header] + \
            user_headers[col_start:col_start+MAX_USERS]

        page_rows = [page_headers]

        for row in rows[1:]:

            page_rows.append(
                [row[0]] +
                row[1+col_start:1+col_start+MAX_USERS]
            )

            table = Table(
                page_rows,
                repeatRows=1,
                hAlign="LEFT",
                colWidths=[35] + [90]*(len(page_headers)-1)
            )

            table.setStyle(TableStyle([

                ("GRID",(0,0),(-1,-1),0.3,colors.grey),

                ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#eeeeee")),

                ("ALIGN",(0,0),(-1,-1),"CENTER"),

                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),

                ("FONTNAME",(0,0),(-1,-1),PDF_JP_FONT),

                ("FONTSIZE",(0,0),(-1,-1),6),

                ("TOPPADDING",(0,0),(-1,-1),1),

                ("BOTTOMPADDING",(0,0),(-1,-1),1),

                ("LEFTPADDING",(0,0),(-1,-1),1),

                ("RIGHTPADDING",(0,0),(-1,-1),1),

            ]))

            story.append(table)
            if col_start + MAX_USERS < len(user_headers):
                story.append(PageBreak())

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title=f"calendar_{mode}_{year}_{month}")
    doc.build(story)
    buffer.seek(0)

    filename = f"calendar_{mode}_{year}_{month}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')


#トップページ
@app.route("/")
def home():
    if "user" not in session:
        return redirect(url_for("login"))
    
    return render_template(
        "home.html", 
        username=session["user"],
        role=session["role"]
    )

#カレンダー表示
@app.route("/calendar/<mode>", methods=["GET"])
def index_get(mode):

    #ログイン確認
    if "user" not in session:
        return redirect(url_for("login"))
    
    username = session["user"]

    #イベントデータ読み込み
    #SAVE_FILE = get_save_file(username, mode)
    events = load_events(username, mode)

    titles = {
        "holiday": "休み希望用",
        "workday": "出勤希望用",
        "childday": "子ども預け希望"
    }
    title = titles.get(mode, "カレンダー")

    color_map = {
        "holiday": "is-danger", #赤
        "workday": "is-info", #青
        "childday": "is-success" #緑
    }
    hero_color = color_map.get(mode, "is-success")
    tag_color = hero_color #前月・翌月タグも同色にする

    #予定表示部分
    bg_map = {
        "holiday": "has-background-danger-light", #赤
        "workday": "has-background-info-light", #青
        "childday": "has-background-success-light" #緑
    }
    event_bg = bg_map.get(mode, "has-background-success-light")

    #デフォルトは今月（パラメータ取得）
    now = datetime.now()
    year = int(request.args.get("year", now.year))
    month = int(request.args.get("month", now.month))
    #入力月の前月を求める
    if month == 1:
        deadline_year = year - 1
        deadline_month = 12

    else:
        deadline_year = year
        deadline_month = month - 1

    #23日0:00で締め切り(つまり22日23:59まで入力可能)
    deadline = datetime(deadline_year, deadline_month, 23, 0, 0)

    is_closed = now >= deadline
    #日曜始まりのカレンダーを作成
    cal = calendar.Calendar(calendar.SUNDAY)
    weeks = cal.monthdayscalendar(year, month)
    #翌月と前月のリンクを作成
    next_year = year
    next_month = month + 1
    if next_month > 12:
        next_month, next_year = 1, year + 1
    prev_year = year
    prev_month = month -1
    if prev_month < 1:
        prev_month, prev_year = 12, year - 1
    next_link = f"?year={next_year}&month={next_month}"
    prev_link = f"?year={prev_year}&month={prev_month}"
    #カレンダーをテンプレートエンジンで表示
    return render_template("index_test2.html",
                           weeknames=list("日月火水木金土"),
                           year=year, month=month,
                           weeks=weeks, events=events,
                           next_link=next_link, prev_link=prev_link,
                           title=title, hero_color=hero_color,
                           tag_color=tag_color, event_bg=event_bg,
                           mode=mode, is_closed=is_closed
                           )

@app.route("/calendar/<mode>", methods=["POST"])
def index_post(mode):

    if "user" not in session:
        return redirect(url_for("login"))
    
    username = session["user"]

    #SAVE_FILE = get_save_file(username, mode)
    events = load_events(username, mode)

    #パラメータを得る
    date = request.form.get("date", "")
    timeS = request.form.get("timeS", "")
    timeE = request.form.get("timeE", "")
    comment = request.form.get("comment", "")

    #削除アクション
    action = request.form.get("action")

    #入力の検証
    i = re.match(r"(\d{4})-(\d{2})-\d{2}", date)
    if not i:
        return "日付形式が不正です。"
    year, month = int(i.group(1)), int(i.group(2))

    #締め切りチェック
    now = datetime.now()

    #入力月の前月を求める
    if month == 1:
        deadline_year = year - 1
        deadline_month = 12

    else:
        deadline_year = year
        deadline_month = month - 1

    #23日0:00で締め切り(つまり22日23:59まで入力可能)
    deadline = datetime(deadline_year, deadline_month, 23, 0, 0)

    if now >= deadline:
        flash("締め切りました。", "warning")
        return redirect(url_for(
            "index_get",
            mode=mode,
            year=year,
            month=month
        ))

    if action == "delete":
        if date in events:
            calendar_repo.delete_event(username, mode, date)
            flash("予定を削除しました", "success")
        else:
            flash("削除対象が見つかりません", "warning")

        return redirect(url_for("index_get", mode=mode, year=year, month=month))

    def valid_time(t):
        return re.fullmatch(r"(0?[7-9]|1[0-9]|2[0-6]):(00|15|30|45)", t)
    
    if not valid_time(timeS) or not valid_time(timeE):
        flash("時刻は7:00~26:00で15分刻みで入力してください。開始時刻と終了時刻両方を入力してください。", "warning")
        return redirect(url_for("index_get", mode=mode, year=year, month=month))
    
    def to_minutes(t):
        h, m = map(int, t.split(":"))
        return h * 60 + m
    
    if to_minutes(timeE) <= to_minutes(timeS):
        flash("終了時刻は開始より後にしてください。", "warning")
        return redirect(url_for("index_get", mode=mode, year=year, month=month))

    calendar_repo.save_event(
        username,
        mode,
        date,
        timeS,
        timeE,
        comment
    )
    return redirect(url_for("index_get", mode=mode, year=year, month=month))

if __name__ == "__main__":
    app.run(debug=True)