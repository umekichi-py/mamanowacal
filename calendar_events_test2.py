import os, json, calendar, re
from datetime import datetime
from flask import Flask, request, render_template, redirect, url_for, session, flash, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Alignment
import io


#admin作成
def init_admin():
    users = load_users()

    if "admin" not in users:
        users["admin"] = {
            "password": generate_password_hash("admin123"),
            "role": "admin"
        }
        save_users(users)
        print("管理者アカウント作成:admin / admin123")

#予定イベントの保存先
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
#json保存先
def get_save_file(username, mode):
    return os.path.join(SCRIPT_DIR, f"calendar_{username}_{mode}.json")
#イベントファイル読み込みの共通関数
def load_events(save_file):
    if os.path.exists(save_file):
        with open(save_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

#ユーザー保存json
def get_user_file():
    return os.path.join(SCRIPT_DIR, "users.json")

def load_users():
    file = get_user_file()
    if os.path.exists(file):
        with open(file, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_users(users):
    with open(get_user_file(), "w", encoding="utf-8") as f:
        json.dump(users, f, ensure_ascii=False, indent=2)

app: Flask = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", os.urandom(24))

#ユーザー登録機能
@app.route("/register", methods=["GET", "POST"])
def register():

#admin以外はとうろくできないように
    if "user" not in session:
        return redirect(url_for("login"))
    
    if session.get("role") != "admin":
        flash("管理者のみ登録できます。", "error")
        return redirect(url_for("home"))

    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        role = request.form.get("role", "user")
        staff_id = request.form.get("staff_id")
        job = request.form.get("job")
        child_name = request.form.get("child_name")

        if not username or not password:
            flash("未入力があります。", "error")
            return render_template("register.html")

        users = load_users()

        if username in users:
            flash("登録済みです。", "error")
            return render_template("register.html")
        
        hashed_pw = generate_password_hash(password)

        users[username] = {
            "password": hashed_pw,
            "role": role,
            "staff_id": staff_id,
            "job": job,
            "child_name": child_name
        }

        save_users(users)

        flash("ユーザー登録が完了しました。", "success")
        return render_template("register.html")
    
    return render_template("register.html")

#ログインページ
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        
        users = load_users()
    
        if username in users and check_password_hash(
            users[username]["password"], password
        ):
            session["user"] = username
            session["role"] = users[username]["role"]
            return redirect(url_for("home"))
        
        flash("ユーザー名またはパスワードが違います。", "error")
        return render_template("login.html")
        
    return render_template("login.html")

#ログアウト
@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

#パスワードの変更機能
@app.route("/change_password", methods=["GET", "POST"])
def change_password():

    if "user" not in session:
        return redirect(url_for("login"))
    
    username = session["user"]

    if request.method == "POST":

        current_pw = request.form.get("current_pw")
        new_pw = request.form.get("new_pw")
        confirm_pw = request.form.get("confirm_pw")

        users = load_users()

        #現在パスワード確認
        if not check_password_hash(users[username]["password"], current_pw):
            flash("現在のパスワードが違います。", "danger")
            return redirect(url_for("change_password"))

        if len(new_pw) < 8:
            flash("パスワードは8文字以上にしてください。")
            return redirect(url_for("change_password"))
        
        if new_pw != confirm_pw:
            flash("新しいパスワードが一致しません。", "danger")
            return redirect(url_for("change_password"))
        
        #更新
        users[username]["password"] = generate_password_hash(new_pw)
        save_users(users)

        flash("パスワードを変更しました。", "success")
        return redirect(url_for("change_password"))
    
    return render_template("change_password.html")

#管理者ページ
@app.route("/admin")
def admin_page():

    if "user" not in session:
        return redirect(url_for("login"))

    if session.get("role") != "admin":
        return "不正な操作です。ページを開きなおしてください。"
    
    users = load_users()

    return render_template("admin.html", users=users)

#ユーザー情報編集ページ
@app.route("/admin/edit_user/<username>", methods=["GET", "POST"])
def edit_user(username):

    if "user" not in session:
        return redirect(url_for("login"))
    
    if session.get("role") != "admin":
        return "不正な操作です"

    users = load_users()

    if username not in users:
        return "ユーザーが存在しません"

    # POST（更新処理）
    if request.method == "POST":
        staff_id = request.form.get("staff_id")
        job = request.form.get("job")
        child_name = request.form.get("child_name")
        role = request.form.get("role")

        if username == "admin" and role != "admin":
            flash("adminの権限は変更できません", "danger")
            return redirect(url_for("admin_page", username=username))

        users[username]["staff_id"] = staff_id
        users[username]["job"] = job
        users[username]["child_name"] = child_name
        users[username]["role"] = role

        save_users(users)

        flash("更新しました", "success")
        return redirect(url_for("admin_page"))

    # GET（表示）
    user = users[username]

    return render_template(
        "edit_user.html",
        username=username,
        user=user
    )

#パスワードのリセット
@app.route("/admin/reset/<username>", methods=["GET", "POST"])
def reset(username):

    if "user" not in session:
        return redirect(url_for("login"))
    
    if session.get("role") != "admin":
        return "不正な操作です。ページを開きなおしてください。"
    
    users = load_users()

    if username not in users:
        flash("ユーザーが存在しません。", "danger")
        return redirect(url_for("reset", username=username))
    
    if request.method == "POST":

        temp_pw = request.form.get("temp_pw")

        if len(temp_pw) < 8:
            flash("パスワードは8文字以上にしてください。", "danger")
            return redirect(url_for("reset", username=username))
        
        users[username]["password"] = generate_password_hash(temp_pw)

        save_users(users)

        flash(f"{username}のパスワードを変更しました。", "success")

        return redirect(url_for("reset", username=username))
    
    return render_template("reset.html", username=username)

#ユーザー削除
@app.route("/admin/delete_user/<username>", methods=["POST"])
def delete_user(username):

    if "user" not in session:
        return redirect(url_for("login"))
    
    if session.get("role") != "admin":
        return "不正な操作です。ページを開きなおしてください。"
    
    users = load_users()

    #admin削除防止
    if username == "admin":
        flash("adminは削除できません。", "danger")
        return redirect(url_for("admin_page"))
    
    if username in users:
        del users[username]
        save_users(users)

        #カレンダーデータ削除
        for mode in ["holiday", "workday", "childday"]:
            file = get_save_file(username, mode)
            if os.path.exists(file):
                os.remove(file)

        flash(f"{username}を削除しました。", "success")

    return redirect(url_for('admin_page'))

#ロール切り替え
@app.route("/admin/toggle_role/<username>", methods=["POST"])
def toggle_role(username):

    if "user" not in session:
        return redirect(url_for("login"))
    
    if session.get("role") != "admin":
        return "不正な操作です。ページを開きなおしてください。"
    
    users = load_users()

    if username not in users:
        flash("ユーザーが存在しません。", "danger")
        return redirect(url_for("admin_page"))
    
    #adminのロール変更防止
    if username == "admin":
        flash("adminの属性は変更できません。", "danger")
        return redirect(url_for("admin_page"))
    
    current_role = users[username]["role"]

    role_order = ["user", "withchild", "childonly", "admin"]

    index = role_order.index(current_role)
    new_role = role_order[(index + 1) % len(role_order)]

    users[username]["role"] = new_role
    save_users(users)

    flash(f"{username}の属性を{new_role}に変更しました。", "success")
    return redirect(url_for("admin_page"))

#一覧カレンダーを選ぶ
@app.route("/admin/calendar")
def admin_calendar():

    if "user" not in session:
        return redirect(url_for("login"))
    
    if session.get("role") != "admin":
        return "不正な操作です。ページを開きなおしてください。"
    
    now = datetime.now()
    default_month = now.strftime("%Y-%m")

    return render_template(
        "admin_calendar_select.html",
        default_month = default_month
        )

#月一覧を作る
@app.route("/admin/calendar/view")
def admin_calendar_view():

    if session.get("role") != "admin":
        return redirect(url_for("login"))
    
    mode = request.args.get("mode")
    month_str = request.args.get("month")

    if not month_str:
        now = datetime.now()
        year = now.year
        month = now.month
    else:
        year, month = map(int, month_str.split("-"))

    days = calendar.monthrange(year, month)[1]

    users = load_users()

    job_order = {"育": 1, "援": 2, "給": 3, "看": 4 }

    users_sorted = sorted(
        users.items(),
        key=lambda x: (
            job_order.get(x[1].get("job") or "", 99),
            x[1].get("staff_id") or "",
            x[0]
        )
    )
                 
    table = {}

    for day in range(1, days +1):
        date = f"{year}-{month:02d}-{day:02d}"
        table[date] = {}

        for username, user in users_sorted:
            file = get_save_file(username, mode)
            if os.path.exists(file):
                with open(file, "r", encoding="utf-8") as f:
                    data = json.load(f)

                table[date][username] = data.get(date, {})

            else:
                table[date][username] = {}

    return render_template(
        "admin_calendar_table.html",
        table = table,
        users = users,
        users_sorted=users_sorted,
        year = year,
        month = month,
        mode = mode
    )

#Excel出力
@app.route("/admin/calendar/export")
def export_calendar():

    def format_time(t):
        if not t:
            return ""
        h, m = t.split(":")
        return f"{int(h)}{m}"

    if session.get("role") != "admin":
        return redirect(url_for("login"))
    
    mode = request.args.get("mode")
    month_str = request.args.get("month")

    if not month_str:
        now = datetime.now()
        year = now.year
        month = now.month

    else:
        year, month = map(int, month_str.split("-"))

    users = load_users()
    days = calendar.monthrange(year, month)[1]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month}"

    #ヘッダー行
    ws.cell(row=1, column=1, value="日付")

    col = 2
    for username in users:
        ws.cell(row=1, column=col, value=username)
        ws.cell(row=1, column=col+1, value="")

        #1ユーザー2列使用
        ws.merge_cells(start_row=1, start_column=col,
                       end_row=1, end_column=col+1)
        col += 2

    #データ
    for day in range(1, days + 1):

        row1 = 2 + (day - 1) * 2 #上段
        row2 = row1 + 1 #下段

        date = f"{year}-{month:02d}-{day:02d}"
        ws.cell(row=row1, column=1, value=date)

        col = 2
        for username in users:
            file = get_save_file(username, mode)

            time_S = ""
            time_E = ""
            comment = ""

            if os.path.exists(file):
                with open(file, "r", encoding="utf-8") as f:
                    data = json.load(f)

                if date in data:
                    e = data[date]
                    time_S = e.get('timeS', '')
                    time_E = e.get('timeE', '')
                    comment = e.get("comment", "")
            
            #上段：開始時刻+コメント
            ws.cell(row=row1, column=col, value=format_time(time_S))
            ws.cell(row=row1, column=col+1, value=comment)

            #下段終了時刻
            ws.cell(row=row2, column=col, value=format_time(time_E))
            
            col += 2

    #メモリに保存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"calendar_{mode}_{year}_{month}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


#トップページ
@app.route("/")
def home():
    if "user" not in session:
        return redirect(url_for("login"))
    
    return render_template(
        "home.html", 
        username=session["user"],
        role=session["role"]
    )

#カレンダー表示
@app.route("/calendar/<mode>", methods=["GET"])
def index_get(mode):

    #ログイン確認
    if "user" not in session:
        return redirect(url_for("login"))
    
    username = session["user"]

    #イベントデータ読み込み
    SAVE_FILE = get_save_file(username, mode)
    events = load_events(SAVE_FILE)

    titles = {
        "holiday": "休み希望用",
        "workday": "出勤希望用",
        "childday": "子ども預け希望"
    }
    title = titles.get(mode, "カレンダー")

    color_map = {
        "holiday": "is-danger", #赤
        "workday": "is-info", #青
        "childday": "is-success" #緑
    }
    hero_color = color_map.get(mode, "is-success")
    tag_color = hero_color #前月・翌月タグも同色にする

    #予定表示部分
    bg_map = {
        "holiday": "has-background-danger-light", #赤
        "workday": "has-background-info-light", #青
        "childday": "has-background-success-light" #緑
    }
    event_bg = bg_map.get(mode, "has-background-success-light")

    #デフォルトは今月（パラメータ取得）
    now = datetime.now()
    year = int(request.args.get("year", now.year))
    month = int(request.args.get("month", now.month))
    deadline = datetime(year, month, 21, 0, 0)
    is_closed = now >= deadline
    #日曜始まりのカレンダーを作成
    cal = calendar.Calendar(calendar.SUNDAY)
    weeks = cal.monthdayscalendar(year, month)
    #翌月と前月のリンクを作成
    next_year = year
    next_month = month + 1
    if next_month > 12:
        next_month, next_year = 1, year + 1
    prev_year = year
    prev_month = month -1
    if prev_month < 1:
        prev_month, prev_year = 12, year - 1
    next_link = f"?year={next_year}&month={next_month}"
    prev_link = f"?year={prev_year}&month={prev_month}"
    #カレンダーをテンプレートエンジンで表示
    return render_template("index_test2.html",
                           weeknames=list("日月火水木金土"),
                           year=year, month=month,
                           weeks=weeks, events=events,
                           next_link=next_link, prev_link=prev_link,
                           title=title, hero_color=hero_color,
                           tag_color=tag_color, event_bg=event_bg,
                           mode=mode, is_closed=is_closed
                           )

@app.route("/calendar/<mode>", methods=["POST"])
def index_post(mode):

    if "user" not in session:
        return redirect(url_for("login"))
    
    username = session["user"]

    SAVE_FILE = get_save_file(username, mode)
    events = load_events(SAVE_FILE)

    #パラメータを得る
    date = request.form.get("date", "")
    timeS = request.form.get("timeS", "")
    timeE = request.form.get("timeE", "")
    comment = request.form.get("comment", "")

    #削除アクション
    action = request.form.get("action")

    #入力の検証
    i = re.match(r"(\d{4})-(\d{2})-\d{2}", date)
    if not i:
        return "日付形式が不正です。"
    year, month = int(i.group(1)), int(i.group(2))

    #締め切りチェック
    now = datetime.now()

    deadline = datetime(year, month, 21, 0, 0)

    if now >= deadline:
        flash("締め切りました。", "danger")
        return redirect(url_for(
            "index_get",
            mode=mode,
            year=year,
            month=month
        ))

    if action == "delete":
        if date in events:
            del events[date]

            with open(SAVE_FILE, "w", encoding="utf-8") as f:
                json.dump(events, f, ensure_ascii=False, indent=2)

            flash("予定を削除しました", "success")
        else:
            flash("削除対象が見つかりません", "warning")

        return redirect(url_for("index_get", mode=mode, year=year, month=month))

    def valid_time(t):
        return re.fullmatch(r"(0?[7-9]|1[0-9]|2[0-6]):(00|15|30|45)", t)
    
    if not valid_time(timeS) or not valid_time(timeE):
        flash("時刻は7:00~26:00で15分刻みで入力してください。開始時刻と終了時刻両方を入力してください。", "warning")
        return redirect(url_for("index_get", mode=mode, year=year, month=month))
    
    def to_minutes(t):
        h, m = map(int, t.split(":"))
        return h * 60 + m
    
    if to_minutes(timeE) <= to_minutes(timeS):
        flash("終了時刻は開始より後にしてください。", "warning")
        return redirect(url_for("index_get", mode=mode, year=year, month=month))

    #イベントを年月日に追加(イベント変数にイベント内容を代入)
    events[date] = {
        "timeS": timeS,
        "timeE": timeE,
        "comment": comment
    }
    with open(SAVE_FILE, "w", encoding="utf-8") as f:
        json.dump(events, f, ensure_ascii=False, indent=2)
    return redirect(url_for("index_get", mode=mode, year=year, month=month))

if __name__ == "__main__":
    init_admin()
    app.run(debug=True)